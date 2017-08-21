
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django_tables2 import RequestConfig
from django.db import IntegrityError
from django.db.models.loading import get_model

from django.contrib.auth.decorators import login_required
from django.template.loader import render_to_string
from django.utils import timezone
from django.middleware.csrf import get_token

from django.template.loader import get_template
from django.template import RequestContext

from .models import *
from .forms import *

from reports.models import *
from .utils import render_to_pdf
import binascii
import re
from django.db import connection, transaction
from django.http import Http404
from mysite.custom_config import *
from reviewer.utils import *
from reports.utils import *

import xlwt
import openpyxl
district_dict = dict(DISTRICT_CHOICES)
#from openpyxl.cell import get_column_letter
# Create your views here.
report_analysis_dict = {}
dict_q = {}
report_templates = { 'Manpower' : 'reviewer/mp.html',
					 'CSC' : 'reviewer/csc.html',
					 'Software' : 'reviewer/software.html',
					 'Hardware' : 'reviewer/hardware.html',
					 'NOFN' : 'reviewer/nofn.html',
					 'SWAN' : 'reviewer/swan.html',
					 'DL' : 'reviewer/digital_literacy.html',
					 'ServiceTrans' : 'reviewer/edistrict_transaction.html',
					 'g2cService' : 'reviewer/g2c_service.html'
					}
pdf_templates = {
					'Manpower' : 'reviewer/manpower_pdf.html',
					'CSC' : 'reviewer/csc_pdf.html',
					 'Software' : 'reviewer/software_pdf.html',
					 'Hardware' : 'reviewer/hardware_pdf.html',
					 'NOFN' : 'reviewer/nofn_pdf.html',
					 'SWAN' : 'reviewer/swan_pdf.html',
					 'DL' : 'reviewer/digital_literacy_pdf.html',
					 'ServiceTrans' : 'reviewer/edistrict_transactioin_pdf.html',
					 'g2cService' : 'reviewer/g2c_service_pdf.html'
	
}
snapshots = {}
@login_required
def analyze_report(request,reportname=None):
	choices = ANALYSIS_CHOICES[reportname]
	if request.method == 'POST':
		form = AnalyzeReportForm(choices,request.POST)
		if form.is_valid():
			choice = form.cleaned_data['analysis_type']
			print choice
			return HttpResponse('Hello')
	
	form = AnalyzeReportForm(choices);
	return render(request,'reviewer/report_analysis.html',{'form':form})

@login_required
def generateReport(request):
	
	if request.method == 'POST':
		form = ReportForm(request.POST)
		null_string = 'No results to display'
		
		if form.is_valid():
			
			district = form.cleaned_data['district']
			district_full_name = district_dict[district]
			report_name = form.cleaned_data['report_name']
			year = form.cleaned_data['year']
			month = form.cleaned_data['month']
			#YYYYMM = form.cleaned_data['YYYYMM']
			YYYYMM = year+month
			report_name = _getCustomReportName(report_name)
			queryset = ReportSnapShot.objects.filter(district=district,YYYYMM=YYYYMM,state='submitted',report_type=report_name)
			if queryset.exists() :
				print "Populating Snapshot from database"
				filtered_result = _massage_content(str(queryset[0].content))
				recipient = queryset[0].owner
				if len(filtered_result) == 0:
					return render(request,'reviewer/empty_results.html',{'null_string':null_string,'report_name':report_name,'district':district_full_name,'YYYYMM':YYYYMM})
				else:
					dict_q[report_name.encode('ascii')] = filtered_result
					#print "updating state of snapshot"
					#queryset.update(state='inprogress')
					return render(request,'reviewer/snapshot.html',{'user':request.user,'filtered_result':filtered_result,'recipient':recipient})
					#return HttpResponse(filtered_result)
			else:
				model = report_name + 'Report'
				model_obj =  get_model('reports', model)
				#queryset = _getSpecficModelQuerySet(report_name)

				#filtered_result = _getFilteredResult(queryset,district,YYYYMM)
				filtered_result = model_obj.objects.filter(district=district,report_month=YYYYMM)
			
				if len(filtered_result) == 0:
					return render(request,'reviewer/empty_results.html',{'null_string':null_string,'report_name':report_name,'district':district_full_name,'YYYYMM':YYYYMM} )

				else:	
					recipient = str(filtered_result[0].user_name)
					dict_q[report_name.encode('ascii')] = filtered_result				
					timestamp  = timezone.now()
					today_date=timestamp.strftime('%Y%m%d')
					#print(filtered_result)
					verbose_report_month = _get_verbose_report_month(YYYYMM)
					context = {'verbose_report_month' : verbose_report_month,'filtered_result':filtered_result,'report_name':report_name,'district_full_name':district_full_name,'YYYYMM':YYYYMM,'today_date':today_date,'district':district,'csrf':1,'include_href':1,'recipient':recipient}
					templ_name = report_templates[report_name]
		
				

					return render(request,templ_name,context )


		else:
			print form.errors
	else:
		form=ReportForm() 

	return render(request,'reviewer/generate_report.html',{'form':form})      


def generatePDF(request):
	district = request.GET.get('district')
	report_month = request.GET.get('period')
	report_name = request.GET.get('report').encode('ascii')
	report_name = _getCustomReportName(report_name)
	model = report_name + 'Report'
	#filtered_result = request.GET.get('report_data')
	#filtered_result = dict_q[report_name]
	#queryset = _getSpecficModelQuerySet(report_name)
	model_obj = get_model('reports',model)
	#filtered_result = _getFilteredResult(queryset,district,report_month)
	verbose_report_month = _get_verbose_report_month(report_month)
	filtered_result =  model_obj.objects.filter(district=district,report_month=report_month)
	pdf_templ = pdf_templates[report_name]
	data = { 'verbose_report_month' : verbose_report_month,'report_month':report_month,'report_name': report_name,'district':district_dict[district], 'filtered_result':filtered_result }
	pdf = render_to_pdf(pdf_templ, data)
	
	#return HttpResponse(dis)

	return HttpResponse(pdf, content_type='application/pdf')



def generateXLSX(request):
	district = request.GET.get('district')
	report_month = request.GET.get('period')
	report_name = request.GET.get('report').encode('ascii')
	report_name = _getCustomReportName(report_name)
	model = report_name + 'Report'
	model_obj = get_model('reports',model)
	filtered_result =  model_obj.objects.filter(district=district,report_month=report_month)
	response =HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

	data = 'attachment; filename='+report_name+'_'+report_month+'_'+district_dict[district]+'.xlsx'
	response['Content-Disposition'] = data
	workbook = openpyxl.Workbook()

	worksheet = workbook.get_active_sheet()
	worksheet.title = report_name+' Report'
	columns = REPORT_HEADERS[report_name]
	row_num = 0
	for col_num in xrange(len(columns)):
		c = worksheet.cell(row=row_num + 1, column=col_num + 1)
		c.value = columns[col_num]
		#c.style.font.bold = True
		# set column width
		#ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

	for i,q in enumerate(filtered_result):
		row_num += 1
		row =_buildListFromQuery(q,report_name)
		for col_num in xrange(len(columns)):
			c = worksheet.cell(row=row_num + 1, column=col_num + 1)
			c.value = row[col_num]
			#c.style.alignment.wrap_text = True
			#print row_num,col_num,row[col_num]
			#worksheet.write(row_num, col_num, row[col_num],font_style)
	workbook.save(response)
	
	return response

def generateXLS(request):
	district = request.GET.get('district')
	report_month = request.GET.get('period')
	report_name = request.GET.get('report').encode('ascii')
	report_name = _getCustomReportName(report_name)
	model = report_name + 'Report'
	model_obj = get_model('reports',model)
	filtered_result =  model_obj.objects.filter(district=district,report_month=report_month)
	response = HttpResponse(content_type='application/ms-excel')
	data = 'attachment; filename='+report_name+'_'+report_month+'_'+district_dict[district]+'.xls'
	response['Content-Disposition'] = data
	workbook = xlwt.Workbook()

	worksheet = workbook.add_sheet(report_name+' Report')

	columns = REPORT_HEADERS[report_name]
	row_num = 0
	for col_num in xrange(len(columns)):
		worksheet.write(row_num,col_num,columns[col_num])

	#data_list = _buildListFromQuery(filtered_result,report_name)
	for i,q in enumerate(filtered_result):
		row_num += 1
		row =_buildListFromQuery(q,report_name)
		for col_num in xrange(len(columns)):
			print row_num,col_num,row[col_num]
			worksheet.write(row_num, col_num, row[col_num])
	workbook.save(response)
	
	return response
	#return HttpResponse('Hello')

	
"""""
Private Routines
"""
def _getModelColumns(report_name):
	model = report_name+'Report'
	model_obj = get_model('reports',model)
	form_obj = MODEL_FORM_BY_MODELNAME[model]
	col_list = []
	exclude_list = list(form_obj._meta.exclude)
	exclude_list.append('id')
	if report_name == 'g2cService' or report_name == 'ServiceTrans':
		exclude_list.remove('associated_line_department')
	model_cols_list = list( model_obj._meta.get_fields() )
	#col_list = [ i for i in model_cols_list if i not in exclude_list ]
	for field in model_cols_list:
		s_field = str(field)
		temp = s_field.split('.')
		col = temp[2]
		if col and col not in exclude_list:
			col_list.append(col)
	
	return col_list


def _buildListFromQuery(q,report_name):
	col_list = _getModelColumns(report_name)
	query_list =[]
	for col in col_list:
		val = getattr(q,col)
		
		if col == 'associated_line_department':
			query_list.append(SERVICE_TO_DEPARTMENT_MAP[q.service_nm])
		else:
			query_list.append(val)

	return query_list


def _convertStringToByte(s):
	lst = []
	for ch in s:
		hv = hex(ord(ch)).replace('0x', '')
		if len(hv) == 1:
			hv = '0'+hv
		lst.append(hv)
	hex_str =''.join( lst )
	bytes = []

	hexStr = ''.join( hex_str.split(" ") )

	for i in range(0, len(hexStr), 2):
		bytes.append( chr( int (hexStr[i:i+2], 16 ) ) )

	return ''.join( bytes )


def _convertByteToStr(byteStr):
	print('F:')
	print (byteStr)
	hex_str = ' '.join( [ "%02X" % ord( x ) for x in byteStr ] )
	
	return  hex_str and chr(atoi(hex_str[:2], base=16)) + toStr(hex_str[2:]) or ''




def _getFilteredResult(queryset,district,yyyymm):

	filtered_result = [ q for i,q in enumerate(queryset) if q.district==district and re.match(yyyymm,q.updated_date) ]

	return filtered_result

def _massage_content(snapshot):
	
	s1 = snapshot.replace("[<ReportSnapShot:","")
	s1 = s1.replace("</html> >]","</html>")
	#s = re.sub(r"^>\]","",s1)
	s = s1.replace(">]","")
	#print(s)

	return s
# def _rawSQLQuery(report_name):
# 	cursor = connection.cursor();
# 	query = "select s.id from reviewer_reportsnapshot s, reviewer_comment c where c.post_id=s.id and c.active=1 and s.report_name=report_name"
def _getCustomReportName(report):
	report_name = report
	if report_name == 'DigitalLiteracy':
		report_name = 'DL' 
	elif report_name == 'G2CServices':
		report_name = 'g2cService' 
	elif report_name == 'eDistrictTrans':
		report_name = 'ServiceTrans'

	return report_name

def _get_verbose_report_month(report_month):
	outstr = ''
	if report_month:
		year = report_month[0:4]
		month = report_month[4:6]
		v_month = custom_config.month[month]
		outstr = v_month + ',' + year


	return outstr